import os
from openpyxl import Workbook
from openpyxl.drawing.image import Image as XLImage
from openpyxl.styles import Alignment, Font, PatternFill
from typing import List, Dict

def generate_excel_from_vacancies(vacancies: List[Dict], logo_path: str, output_path: str):
    wb = Workbook()
    ws = wb.active
    ws.title = "Вакансии"

    headers = ["Должность", "Доход", "Локация", "Работодатель", "Условия", "Контакты", "Источник"]
    ws.append(headers)

    header_font = Font(bold=True, color="FFFFFF")
    fill = PatternFill("solid", fgColor="2F75B5")
    align = Alignment(horizontal="center", vertical="center", wrap_text=True)

    for col_num, col in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num)
        cell.font = header_font
        cell.fill = fill
        cell.alignment = align
        ws.column_dimensions[cell.column_letter].width = 30

    for v in vacancies:
        ws.append([
            v.get("title", ""),
            v.get("salary", ""),
            v.get("location", ""),
            v.get("employer", ""),
            v.get("conditions", ""),
            v.get("contact", ""),
            v.get("source", ""),
        ])

    if os.path.exists(logo_path):
        img = XLImage(logo_path)
        img.anchor = 'H2'
        ws.add_image(img)

    wb.save(output_path)
